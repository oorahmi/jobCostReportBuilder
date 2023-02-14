import openpyxl

import sys
from openpyxl.styles import Border, Side
from copy import copy

def set_border(ws, cell_range):
# https://stackoverflow.com/questions/34520764/apply-border-to-range-of-cells-using-openpyxl
    rows = ws[cell_range]
    side = Side(border_style='thick', color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border

# draw a line, am just removing all side and top capability from set_border, this may have unexpected behavior
# draws expected line at row - 1
def draw_line(ws, cell_range):
    rows = ws[cell_range]
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                bottom=cell.border.bottom
            )
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_y == 0 or pos_y == max_y:
                cell.border = border



# cant copy sheet from one workbook to another without a deep level copy
# https://stackoverflow.com/questions/42344041/how-to-copy-worksheet-from-one-workbook-to-another-one-using-openpyxl
def copySheet(source_sheet, target_sheet):

    def copy_sheet_attributes(source_sheet, target_sheet):
        target_sheet.sheet_format = copy(source_sheet.sheet_format)
        target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
        target_sheet.merged_cells = copy(source_sheet.merged_cells)
        target_sheet.page_margins = copy(source_sheet.page_margins)
        target_sheet.freeze_panes = copy(source_sheet.freeze_panes)
        # set row dimensions
        # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
        for rn in range(len(source_sheet.row_dimensions)):
            target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

        if source_sheet.sheet_format.defaultColWidth is None:
            print('Unable to copy default column wide')
        else:
            target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

        # set specific column width and hidden property
        # we cannot copy the entire column_dimensions attribute so we copy selected attributes
        for key, value in source_sheet.column_dimensions.items():
            target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
            target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
            target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
            target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

    def copy_cells(source_sheet, target_sheet):
        for (row, col), source_cell in source_sheet._cells.items():
            target_cell = target_sheet.cell(column=col, row=row)
            target_cell._value = source_cell._value
            target_cell.data_type = source_cell.data_type
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

            if source_cell.hyperlink:
                target_cell._hyperlink = copy(source_cell.hyperlink)

            if source_cell.comment:
                target_cell.comment = copy(source_cell.comment)

    copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)
        

def progressbar(it, prefix="", size=60, out=sys.stdout): # Python3.3+
    count = len(it)
    def show(j):
        x = int(size*j/count)
        print("{}[{}{}] {}/{}".format(prefix, "#"*x, "."*(size-x), j, count), 
                end='\r', file=out, flush=True)
    show(0)
    for i, item in enumerate(it):
        yield item
        show(i+1)
    print("\n", flush=True, file=out)